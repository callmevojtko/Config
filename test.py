import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

def create_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet
    return wb

def add_sheet(wb, title):
    return wb.create_sheet(title=title)

def style_header(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(bottom=Side(style='thin'))

def add_data_to_sheet(ws, data):
    for row in data:
        ws.append(row)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

def main():
    wb = create_workbook()

    # Configuration Page
    ws = add_sheet(wb, "Configuration Page")
    ws.append(["Configuration Page"])
    ws.append(["HP LaserJet 600 M602"])
    ws.append(["Page 1"])

    # Device Information
    ws = add_sheet(wb, "Device Information")
    data = [
        ["Key", "Value"],
        ["Product Name", "HP LaserJet 600 M602"],
        ["Device Name", "HP LaserJet 600 M602"],
        ["Model Number", "CE991A"],
        ["DC Controller Version", "5.121"],
        ["Duplex Unit Version", "2.101"],
        ["Product Serial Number", "CNBCD4M10P"],
        ["Formatter Number", "TD20GVV"],
        ["Firmware Bundle Version", "3.8"],
        ["Firmware Revision", "2308209_000586"],
        ["Firmware Datecode", "20160820"],
        ["Service ID", "22262"],
        ["HP FutureSmart Level", "HP FutureSmart 2"],
        ["Pages Since Last Maintenance", "2611"],
        ["PS Wait Time-out", "300 seconds"],
        ["Engine Cycles", "2615"],
        ["*Not weighted for billing", ""]
    ]
    add_data_to_sheet(ws, data)
    style_header(ws['A1'])
    style_header(ws['B1'])

    # Memory
    ws = add_sheet(wb, "Memory")
    data = [
        ["Key", "Value"],
        ["Total RAM", "512 MB"]
    ]
    add_data_to_sheet(ws, data)
    style_header(ws['A1'])
    style_header(ws['B1'])

    # Event Log
    ws = add_sheet(wb, "Event Log")
    data = [
        ["Key", "Value"],
        ["Number of Entries in Use", "66"],
        ["Three Most Recent Entries", ""],
        ["Number", "Cycles", "Event"],
        ["66", "2599", "13.82,01"],
        ["65", "2598", "13.82.01"],
        ["54", "2559", "41.03.01"]
    ]
    add_data_to_sheet(ws, data)
    style_header(ws['A1'])
    style_header(ws['B1'])
    style_header(ws['A4'])
    style_header(ws['B4'])
    style_header(ws['C4'])

    # Installed Personalities and Options
    ws = add_sheet(wb, "Installed Personalities")
    data = [
        ["Component", "Version"],
        ["PCL", "20010402"],
        ["PCLXL", "20010402"],
        ["POSTSCRIPT", "20010402"],
        ["PDF", "20130901"],
        ["AirPrint", "2040201"],
        ["PWGRASTER", "2040201"],
        ["TIFF", "20150522"],
        ["Internal Disk", "Hard Disk: Enabled"],
        ["Serial Number", "120946313901"],
        ["Model", "SanDisk SSD P4 468"],
        ["Capacity", "3825 MB"],
        ["Hard Disk Encryption Status", "Disk cannot be encrypted"],
        ["Embedded HP JetDirect J8028E", "296-TechEd-P1"]
    ]
    add_data_to_sheet(ws, data)
    style_header(ws['A1'])
    style_header(ws['B1'])

    # HP Web Services
    ws = add_sheet(wb, "HP Web Services")
    data = [
        ["Service", "Status"],
        ["HP Web Services", "Disabled"],
        ["ePrint", "Disabled"]
    ]
    add_data_to_sheet(ws, data)
    style_header(ws['A1'])
    style_header(ws['B1'])

    # Security
    ws = add_sheet(wb, "Security")
    data = [
        ["Key", "Value"],
        ["Hard Disk Encryption Status", ""],
        ["Internal Disk", "Disk cannot be encrypted"],
        ["Job Data Encryption Status", "Encrypted (AES-128)"],
        ["Job Data Persistence", "Volatile (not persistent)"],
        ["File Erase Mode", "Non-Secure Fast Erase (No overwrite)"],
        ["Control Panel Password", "Disabled"],
        ["Support Key", "J52A-3FUP-8LCR"],
        ["Host USB plug and play", "Enabled"],
        ["Device USB", "Enabled"],
        ["Whitelisting", "Present"],
        ["Runtime Intrusion Detection", "Present"]
    ]
    add_data_to_sheet(ws, data)
    style_header(ws['A1'])
    style_header(ws['B1'])

    # Paper Trays and Options
    ws = add_sheet(wb, "Paper Trays and Options")
    data = [
        ["Key", "Value"],
        ["Default Paper Size", "Letter"],
        ["Tray 1 Size", "Envelope #10"],
        ["Tray 1 Type", "Envelope"],
        ["Tray 2 Size", "Letter"],
        ["Tray 2 Type", "Plain"],
        ["Duplex Unit", ""],
        ["1", "Cutput Bin 1. 500 Sheets, Standard bin (correct order)"],
        ["2", "Output Bin 2, 100 Sheets, Rear bin (straightest path)"]
    ]
    add_data_to_sheet(ws, data)
    style_header(ws['A1'])
    style_header(ws['B1'])

    # Additional Info
    ws = add_sheet(wb, "Additional Info")
    data = [
        ["Key", "Value"],
        ["Fraser ID", "A1234"],
        ["Building Name", "RTLC"],
        ["ROOM #", "101"],
        ["Floor", "1st Floor"],
        ["Location", "Blake office"]
    ]
    add_data_to_sheet(ws, data)
    style_header(ws['A1'])
    style_header(ws['B1'])

    wb.save("printer_config_organized.xlsx")
    print("Excel file 'printer_config_organized.xlsx' has been created successfully.")

if __name__ == "__main__":
    main()