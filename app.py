import cv2
import pytesseract
from openpyxl import Workbook
import logging
import re
from pdf2image import convert_from_path
import numpy as np
import os

# pytesseract path
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def preprocess_image(image):
    logging.info("Preprocessing image")
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    denoised = cv2.fastNlMeansDenoising(gray, None, 10, 7, 21)
    thresh = cv2.threshold(denoised, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
    
    # Add dilation to connect broken characters
    kernel = np.ones((1,1), np.uint8)
    dilated = cv2.dilate(thresh, kernel, iterations=1)
    
    return dilated

def extract_text_from_image(image):
    logging.info("Extracting text from image")
    custom_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789:.-_() '
    text = pytesseract.image_to_string(image, config=custom_config)
    if not text.strip():
        raise ValueError("No text extracted from the image")
    return text

def process_pdf(pdf_path):
    logging.info(f"Processing PDF: {pdf_path}")
    pages = convert_from_path(pdf_path, dpi=300)  # Increase DPI for better quality
    all_text = ""
    for i, page in enumerate(pages):
        logging.info(f"Processing page {i+1}")
        img_np = np.array(page)
        img_cv = cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR)
        preprocessed = preprocess_image(img_cv)
        text = extract_text_from_image(preprocessed)
        all_text += text + "\n\n"
    return all_text

def parse_config_data(text):
    logging.info("Parsing configuration data")
    sections = {}
    current_section = None
    lines = text.split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Check if this line is a header (all uppercase or title case)
        if line.isupper() or line.istitle():
            current_section = line
            sections[current_section] = []
        elif current_section and ':' in line:
            key, value = line.split(':', 1)
            sections[current_section].append([key.strip(), value.strip()])
        elif current_section:
            # If there's no colon, add the entire line as a value
            sections[current_section].append(['', line])
    
    return sections

def save_to_excel(data, output_path):
    logging.info(f"Saving data to Excel: {output_path}")
    wb = Workbook()
    ws = wb.active
    ws.title = "Configuration Data"

    # Write headers
    headers = list(data.keys())
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Write data
    max_rows = max(len(section) for section in data.values())
    for row in range(2, max_rows + 2):
        for col, header in enumerate(headers, start=1):
            if row - 2 < len(data[header]):
                value = ' - '.join(data[header][row-2])
                ws.cell(row=row, column=col, value=value)

    try:
        wb.save(output_path)
    except Exception as e:
        logging.error(f"Failed to save Excel file: {e}")
        raise

def main():
    # Use the correct relative path
    pdf_path = os.path.abspath("scan_bbossler_2024-06-20-15-22-11.pdf")
    output_path = "printer_config.xlsx"
    
    # Add debugging information
    logging.info(f"Current working directory: {os.getcwd()}")
    logging.info(f"Attempting to process PDF at: {pdf_path}")
    logging.info(f"File exists: {os.path.exists(pdf_path)}")
    
    try:
        extracted_text = process_pdf(pdf_path)
        config_data = parse_config_data(extracted_text)
        save_to_excel(config_data, output_path)
        logging.info(f"Configuration data has been extracted and saved to {output_path}")
    except Exception as e:
        logging.error(f"An error occurred: {e}")

if __name__ == "__main__":
    main()